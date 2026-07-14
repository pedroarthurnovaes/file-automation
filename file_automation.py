from __future__ import annotations

import argparse
import json
import logging
import re
import shutil
import sys
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


SUPPORTED_EXTENSIONS = (".xlsx", ".xlsm")
APPLY_CONFIRMATION = "REPLACE"


class GuardError(RuntimeError):
    """A safe, user-actionable failure."""


@dataclass(frozen=True)
class Config:
    active_dir: Path
    incoming_dir: Path
    archive_dir: Path
    active_stem: str
    incoming_stem: str
    date_column: str
    require_incoming_not_older: bool = True
    overwrite_existing_archive: bool = False


@dataclass(frozen=True)
class UpdatePlan:
    active_file: Path
    archive_target: Path
    incoming_file: Path
    replacement_target: Path
    active_date: date
    incoming_date: date


def normalize_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def normalize_stem(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip())


def resolve_config_path(value: str, config_dir: Path) -> Path:
    candidate = Path(value).expanduser()
    return candidate if candidate.is_absolute() else (config_dir / candidate).resolve()


def load_config(path: Path) -> Config:
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except FileNotFoundError as exc:
        raise GuardError(f"Configuration file not found: {path}") from exc
    except json.JSONDecodeError as exc:
        raise GuardError(f"Configuration is not valid JSON: {exc}") from exc

    required = (
        "active_dir",
        "incoming_dir",
        "archive_dir",
        "active_stem",
        "incoming_stem",
        "date_column",
    )
    missing = [key for key in required if not raw.get(key)]
    if missing:
        raise GuardError(f"Configuration is missing: {', '.join(missing)}")

    base = path.resolve().parent
    return Config(
        active_dir=resolve_config_path(raw["active_dir"], base),
        incoming_dir=resolve_config_path(raw["incoming_dir"], base),
        archive_dir=resolve_config_path(raw["archive_dir"], base),
        active_stem=str(raw["active_stem"]),
        incoming_stem=str(raw["incoming_stem"]),
        date_column=str(raw["date_column"]),
        require_incoming_not_older=bool(raw.get("require_incoming_not_older", True)),
        overwrite_existing_archive=bool(raw.get("overwrite_existing_archive", False)),
    )


def parse_date(value: object) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date()
        except (TypeError, ValueError, OverflowError):
            return None
    if isinstance(value, str):
        text = value.strip()
        for fmt in (
            "%Y-%m-%d",
            "%m/%d/%Y",
            "%d/%m/%Y",
            "%Y/%m/%d",
            "%m-%d-%Y",
            "%d-%m-%Y",
        ):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    return None


def matching_files(folder: Path, stems: Iterable[str]) -> list[Path]:
    if not folder.is_dir():
        raise GuardError(f"Folder not found: {folder}")
    expected = {normalize_stem(stem).casefold() for stem in stems}
    matches = [
        path
        for path in folder.iterdir()
        if path.is_file()
        and path.suffix.casefold() in SUPPORTED_EXTENSIONS
        and normalize_stem(path.stem).casefold() in expected
    ]
    return sorted(matches)


def find_active_file(config: Config) -> Path:
    matches = matching_files(config.active_dir, [config.active_stem])
    if len(matches) != 1:
        description = "none" if not matches else ", ".join(path.name for path in matches)
        raise GuardError(f"Expected exactly one active workbook; found {description}")
    return matches[0]


def find_incoming_file(config: Config) -> Path:
    if not config.incoming_dir.is_dir():
        raise GuardError(f"Folder not found: {config.incoming_dir}")
    pattern = re.compile(
        rf"^{re.escape(normalize_stem(config.incoming_stem))}(?: \(\d+\))?$",
        re.IGNORECASE,
    )
    candidates = [
        path
        for path in config.incoming_dir.iterdir()
        if path.is_file()
        and path.suffix.casefold() in SUPPORTED_EXTENSIONS
        and pattern.match(normalize_stem(path.stem))
    ]
    if not candidates:
        raise GuardError(
            f"No incoming workbook matching '{config.incoming_stem}' in {config.incoming_dir}"
        )
    return max(candidates, key=lambda path: path.stat().st_mtime)


def max_column_date(workbook_path: Path, column_name: str) -> date:
    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except PermissionError as exc:
        raise GuardError(
            f"Cannot read '{workbook_path}'. Close it in Excel and wait for sync to finish."
        ) from exc
    except OSError as exc:
        raise GuardError(f"Cannot open workbook '{workbook_path}': {exc}") from exc

    expected = normalize_text(column_name)
    try:
        for sheet in workbook.worksheets:
            header_row = None
            column_index = None
            for row_number, row in enumerate(
                sheet.iter_rows(max_row=20, values_only=True), start=1
            ):
                headers = [normalize_text(value) for value in row]
                if expected in headers:
                    header_row = row_number
                    column_index = headers.index(expected)
                    break
            if header_row is None or column_index is None:
                continue

            parsed_dates = []
            for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                if column_index < len(row):
                    parsed = parse_date(row[column_index])
                    if parsed:
                        parsed_dates.append(parsed)
            if parsed_dates:
                return max(parsed_dates)
    finally:
        workbook.close()

    raise GuardError(
        f"Column '{column_name}' was not found with usable dates in {workbook_path.name}"
    )


def build_plan(config: Config) -> UpdatePlan:
    active_file = find_active_file(config)
    incoming_file = find_incoming_file(config)
    active_date = max_column_date(active_file, config.date_column)
    incoming_date = max_column_date(incoming_file, config.date_column)

    if config.require_incoming_not_older and incoming_date < active_date:
        raise GuardError(
            f"Incoming workbook is older ({incoming_date}) than the active workbook ({active_date})."
        )

    archive_target = (
        config.archive_dir
        / f"{config.active_stem}_{active_date:%Y%m%d}{active_file.suffix.casefold()}"
    )
    if archive_target.exists() and not config.overwrite_existing_archive:
        raise GuardError(
            f"Archive already exists: {archive_target}. Refusing to overwrite it."
        )

    return UpdatePlan(
        active_file=active_file,
        archive_target=archive_target,
        incoming_file=incoming_file,
        replacement_target=config.active_dir
        / f"{config.active_stem}{incoming_file.suffix.casefold()}",
        active_date=active_date,
        incoming_date=incoming_date,
    )


def describe_plan(plan: UpdatePlan) -> None:
    logging.info("Active workbook: %s", plan.active_file)
    logging.info("Active data date: %s", plan.active_date)
    logging.info("Archive target: %s", plan.archive_target)
    logging.info("Incoming workbook: %s", plan.incoming_file)
    logging.info("Incoming data date: %s", plan.incoming_date)
    logging.info("Replacement target: %s", plan.replacement_target)


def apply_plan(plan: UpdatePlan, overwrite_archive: bool = False) -> None:
    plan.archive_target.parent.mkdir(parents=True, exist_ok=True)
    archived = False
    try:
        if plan.archive_target.exists():
            if not overwrite_archive:
                raise GuardError(f"Archive already exists: {plan.archive_target}")
            plan.archive_target.unlink()
        shutil.move(str(plan.active_file), str(plan.archive_target))
        archived = True
        shutil.move(str(plan.incoming_file), str(plan.replacement_target))
    except (OSError, shutil.Error) as exc:
        if archived and plan.archive_target.exists() and not plan.active_file.exists():
            try:
                shutil.move(str(plan.archive_target), str(plan.active_file))
            except (OSError, shutil.Error) as rollback_exc:
                raise GuardError(
                    "Replacement failed and rollback also failed. "
                    f"Archive remains at '{plan.archive_target}'. Rollback error: {rollback_exc}"
                ) from exc
        raise GuardError(f"No replacement was completed: {exc}") from exc


def configure_logging(log_dir: Path) -> Path:
    log_dir.mkdir(parents=True, exist_ok=True)
    log_path = log_dir / f"file_automation_{datetime.now():%Y%m%d_%H%M%S}.log"
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        handlers=[
            logging.FileHandler(log_path, encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ],
        force=True,
    )
    return log_path


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Validate, archive, and safely replace a workbook export."
    )
    parser.add_argument("--config", type=Path, default=Path("config.json"))
    parser.add_argument("--apply", action="store_true", help="Move files after validation.")
    parser.add_argument(
        "--confirm",
        help=f"Required with --apply; enter {APPLY_CONFIRMATION}.",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    log_path = configure_logging(Path(__file__).resolve().parent / "logs")
    try:
        config = load_config(args.config)
        plan = build_plan(config)
        describe_plan(plan)
        if not args.apply:
            logging.info("DRY RUN complete. No files were changed.")
            return 0
        if args.confirm != APPLY_CONFIRMATION:
            raise GuardError(
                f"Apply mode requires '--confirm {APPLY_CONFIRMATION}'. No files were changed."
            )
        apply_plan(plan, overwrite_archive=config.overwrite_existing_archive)
        logging.info("APPLY complete. The active workbook was safely replaced.")
        return 0
    except GuardError as exc:
        logging.error("%s", exc)
        return 1
    finally:
        print(f"Log file: {log_path}")


if __name__ == "__main__":
    raise SystemExit(main())
